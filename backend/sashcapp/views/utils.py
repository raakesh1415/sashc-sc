# ─── Shared grade constants ────────────────────────────────────────────────────

VALID_EXAMS = ['TOV', 'AR1', 'AR2', 'ETR']
GRADE_KEYS  = ['A+', 'A', 'A-', 'B+', 'B', 'C+', 'C', 'D', 'E', 'G']
GP_WEIGHTS = {'A+': 0, 'A': 1, 'A-': 2, 'B+': 3, 'B': 4, 'C+': 5, 'C': 6, 'D': 7, 'E': 8, 'G': 9}


def calc_gp(grade_counts, total):
    """Return GP score given a dict of {grade: count} and a total count."""
    if total == 0:
        return 0
    numerator = sum(GP_WEIGHTS.get(g, 0) * grade_counts.get(g, 0) for g in GP_WEIGHTS)
    return numerator / total


def empty_grade_counts(include_th=False):
    """Return a zeroed grade-count dict, optionally including 'TH'."""
    counts = {g: 0 for g in GRADE_KEYS}
    if include_th:
        counts['TH'] = 0
    return counts


def lulus_status(bm_mark, sej_mark):
    """Return 'LULUS' if both BM and SEJ marks are >= 40, else 'GAGAL'."""
    if bm_mark is not None and sej_mark is not None and bm_mark >= 40 and sej_mark >= 40:
        return 'LULUS'
    return 'GAGAL'


# ─── Bulk-upload shared helpers ───────────────────────────────────────────────

ALLOWED_EMAIL_DOMAINS = {'moe-dl.edu.my', 'gmail.com'}

# Known template signatures — unique column sets per template type
_UPLOAD_SIGNATURES = {
    'student':  {'Subjects'},                       # unique to student user template
    'teacher':  {'Roles', 'Subject Classes'},        # unique to teacher user template
    'class':    {'Class Name'},                     # class template
    'subject':  {'Subject Name', 'Subject Code'},   # subject template
    'tov':      {'Student Name', 'Class Name'},     # TOV pivot marks template
    'marks':    {'Student Name', 'Subject Name', 'Mark'},  # AR1/AR2/ETR flat template
}

_UPLOAD_LABELS = {
    'student': 'PELAJAR',
    'teacher': 'GURU',
    'class':   'KELAS',
    'subject': 'MATA PELAJARAN',
    'tov':     'MARKAH TOV',
    'marks':   'MARKAH (AR/ETR)',
}


def detect_wrong_template(fieldnames_set, expected_template, required_cols=None):
    """
    Check if the uploaded file matches a known template OTHER than the expected one,
    or is completely unrecognized.

    Returns an error payload dict in three cases:
      1. File matches a known wrong template  → specific "wrong template" message
      2. File is completely unrecognized      → "template tidak dikenali" message
      3. File is missing required columns     → "template tidak lengkap" message (only if required_cols given)
    Returns None if the file looks correct.
    """
    expected_label = _UPLOAD_LABELS.get(expected_template, expected_template.upper())

    # ── 1. Check for known wrong template ────────────────────────────────────
    for tpl_key, sig in _UPLOAD_SIGNATURES.items():
        if tpl_key == expected_template:
            continue
        if not (sig <= fieldnames_set):
            continue
        # 'class' template shares 'Class Name' with marks templates — only flag it
        # when there are NO other user/marks fields present.
        if tpl_key == 'class' and any(
            f in fieldnames_set for f in ('Email', 'Name', 'Student Name', 'Subject Name', 'Mark', 'Roles')
        ):
            continue
        # 'tov' shares 'Student Name'+'Class Name' with marks template — skip if 'Mark' also present
        if tpl_key == 'tov' and 'Mark' in fieldnames_set:
            continue
        detected_label = _UPLOAD_LABELS[tpl_key]
        return {
            'message': f'Template yang salah! Sila gunakan template {expected_label}.',
            'errorCount': 1, 'successCount': 0,
            'errors': [f'Fail yang dimuat naik kelihatan seperti template {detected_label}.'],
        }

    # ── 2. Check required columns (missing = incomplete or unrecognized) ─────
    if required_cols:
        missing = required_cols - fieldnames_set
        if missing:
            # If NO expected columns are present at all → completely unrecognized file
            if not (required_cols & fieldnames_set):
                return {
                    'message': f'Template tidak dikenali! Sila muat turun dan gunakan template {expected_label} yang disediakan.',
                    'errorCount': 1, 'successCount': 0,
                    'errors': ['Fail yang dimuat naik bukan template yang digunakan dalam sistem ini.'],
                }
            # Some expected columns exist but some are missing → incomplete template
            return {
                'message': f'Template tidak lengkap! Kolum yang tiada: {", ".join(sorted(missing))}',
                'errorCount': 1, 'successCount': 0,
                'errors': [f'Kolum diperlukan: {", ".join(sorted(missing))}'],
            }

    return None


def get_subject_by_name_or_code(value, session_year):
    """Look up a Subject by subjectName OR subjectCode (case-insensitive), scoped to session_year."""
    from django.db.models import Q
    from ..models import Subject
    qs = Subject.objects.filter(Q(subjectName__iexact=value) | Q(subjectCode__iexact=value))
    if session_year:
        qs = qs.filter(year=session_year)
    return qs.get()


def apply_mark(headcount, mark_field, grade_field, mark_str):
    """
    Validate mark_str and apply to headcount fields.
    Returns (ok: bool, error_msg: str | None).
    Empty string → skip (returns True, None without saving).
    """
    if mark_str == '':
        return True, None
    if mark_str.upper() == 'TH':
        setattr(headcount, mark_field, None)
        setattr(headcount, grade_field, 'TH')
        return True, None
    try:
        mark_value = float(mark_str)
    except ValueError:
        return False, f"Markah '{mark_str}' tidak sah. Masukkan nombor 0-100 atau 'TH'"
    if mark_value < 0 or mark_value > 100:
        return False, f"Markah '{mark_str}' mestilah antara 0 dan 100"
    setattr(headcount, mark_field, mark_value)
    setattr(headcount, grade_field, _grade(mark_value))
    return True, None


def build_upload_caches(session_year, subject_filter_ids=None, subject_key='name'):
    """
    Build in-memory caches for bulk CSV upload processing.

    Avoids N+1 queries by pre-loading subjects, students, and headcounts
    into dicts before iterating CSV rows.

    Args:
        session_year: int or None — filters all lookups to this year.
        subject_filter_ids: iterable of subject PKs to restrict student/headcount
                            queries (pass None to load all students).
        subject_key: 'name' (default) to index by subjectName (AR1/AR2/ETR),
                     'code' to index by subjectCode (TOV).

    Returns a dict with keys:
        subject_cache      — {key_lower: Subject}
        subject_duplicates — set of ambiguous subject keys
        student_cache      — {(name_lower, class_lower): User}
        student_duplicates — set of ambiguous student keys
        hc_cache           — {(user_id_str, subject_id_str): Headcount}
    """
    from ..models import Headcount, Subject, User

    # Subjects
    subj_qs = Subject.objects.all()
    if subject_filter_ids is not None:
        subj_qs = subj_qs.filter(subjectID__in=subject_filter_ids)
    if session_year:
        subj_qs = subj_qs.filter(year=session_year)

    subject_cache: dict = {}
    subject_duplicates: set = set()
    for s in subj_qs:
        raw = s.subjectCode if subject_key == 'code' else s.subjectName
        key = raw.lower() if raw else None
        if not key:
            continue
        if key in subject_cache:
            subject_duplicates.add(key)
        subject_cache[key] = s

    # Students
    student_qs = User.objects.filter(studentID__isnull=False).select_related('studentID__classID')
    if subject_filter_ids is not None:
        student_qs = student_qs.filter(
            studentID__enrollSubjectID__enrollments__subjectID__in=subject_filter_ids
        ).distinct()
    if session_year:
        student_qs = student_qs.filter(year=session_year)

    student_cache: dict = {}
    student_duplicates: set = set()
    for u in student_qs:
        cn = (u.studentID.classID.className or '') if u.studentID and u.studentID.classID else ''
        key = (u.name.lower(), cn.lower())
        if key in student_cache:
            student_duplicates.add(key)
        student_cache[key] = u

    # Headcounts
    all_user_ids = [u.pk for u in student_cache.values()]
    hc_qs = Headcount.objects.filter(userID__in=all_user_ids)
    if subject_filter_ids is not None:
        hc_qs = hc_qs.filter(subjectID__in=subject_filter_ids)
    hc_cache: dict = {
        (str(hc.userID_id), str(hc.subjectID_id)): hc
        for hc in hc_qs
    }

    return {
        'subject_cache':      subject_cache,
        'subject_duplicates': subject_duplicates,
        'student_cache':      student_cache,
        'student_duplicates': student_duplicates,
        'hc_cache':           hc_cache,
    }


def get_active_year(request):
    """
    Read the active academic year from ?year= query param.
    Returns int (2000–2100) or None (None means no filter is applied).
    """
    year_str = request.query_params.get('year')
    if year_str:
        try:
            year = int(year_str)
            if 2000 <= year <= 2100:
                return year
        except (ValueError, TypeError):
            pass
    return None


# ─── XLSX template helpers ────────────────────────────────────────────────────

def make_xlsx_response(filename, headers, rows, readonly_cols=None, number_cols=None):
    """
    Build an xlsx HttpResponse with coloured headers.
    headers       — list of str
    rows          — list of lists (data rows)
    readonly_cols — set of 0-based column indices to shade light grey (not editable)
    number_cols   — set of 0-based column indices to format as numbers (0.##)
    Header row is always light yellow; readonly_cols data cells are light grey.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment
    from django.http import HttpResponse

    readonly_cols = readonly_cols or set()
    number_cols   = number_cols   or set()
    HEADER_FILL   = PatternFill(fill_type='solid', fgColor='FFFF99')
    READONLY_FILL = PatternFill(fill_type='solid', fgColor='ECECEC')
    _side         = Side(style='thin', color='AAAAAA')
    THIN_BORDER   = Border(left=_side, right=_side, top=_side, bottom=_side)
    LEFT_ALIGN    = Alignment(horizontal='left')

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = HEADER_FILL
        cell.border    = THIN_BORDER
        cell.alignment = LEFT_ALIGN

    for row_data in rows:
        ws.append(row_data)
        row_num = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = LEFT_ALIGN
        for col_0 in readonly_cols:
            cell = ws.cell(row=row_num, column=col_0 + 1)
            cell.fill   = READONLY_FILL
            cell.border = THIN_BORDER
        for col_0 in number_cols:
            cell = ws.cell(row=row_num, column=col_0 + 1)
            if cell.value not in (None, ''):
                try:
                    val = float(cell.value)
                    cell.value = int(val) if val == int(val) else val
                except (ValueError, TypeError):
                    pass

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def read_upload_rows(file_obj):
    """
    Read a CSV or XLSX upload file and return a list of row dicts.
    Strips whitespace from all keys and values.
    Skips fully-empty rows (xlsx artefact).
    """
    import csv, io as _io
    name = getattr(file_obj, 'name', '')
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return []
        headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        result = []
        for row in all_rows[1:]:
            if all(v is None for v in row):
                continue
            d = {headers[i]: (str(row[i]).strip() if row[i] is not None else '')
                 for i in range(len(headers))}
            result.append(d)
        return result
    else:
        content = file_obj.read().decode('utf-8-sig')
        return list(csv.DictReader(_io.StringIO(content)))


def _grade(mark):
    """Convert a numeric mark (0–100) to a grade string."""
    v = float(mark)
    if v >= 90: return 'A+'
    if v >= 80: return 'A'
    if v >= 70: return 'A-'
    if v >= 65: return 'B+'
    if v >= 60: return 'B'
    if v >= 55: return 'C+'
    if v >= 50: return 'C'
    if v >= 45: return 'D'
    if v >= 40: return 'E'
    return 'G'
